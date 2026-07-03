from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.config import Settings
from app.models import BudgetResponse, DesignPlan, ProductMatch


class ExcelService:
    def __init__(self, settings: Settings):
        self.settings = settings

    def export(self, plan: DesignPlan, matches: list[ProductMatch], budget: BudgetResponse) -> Path:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "采购清单"
        headers = ["单品", "材质", "尺寸", "搜索关键词", "商品名", "券后价", "原价", "佣金", "销量", "店铺", "来源", "商品图", "直达链接"]
        sheet.append(headers)
        for match in matches:
            for product in match.products:
                sheet.append(
                    [
                        match.design_item.name,
                        match.design_item.material,
                        match.design_item.size,
                        match.design_item.taobao_keyword,
                        product.title,
                        product.final_price,
                        product.original_price or product.price,
                        f"{product.commission_rate:.2f}%",
                        product.sales,
                        product.shop_name,
                        product.source,
                        product.image_url or "",
                        product.item_url,
                    ]
                )
        budget_sheet = workbook.create_sheet("预算汇总")
        budget_sheet.append(["版本", "目标人群", "总价", "说明"])
        budget_sheet.append([budget.low_plan.name, budget.low_plan.target_user, budget.low_plan.total_price, budget.low_plan.note])
        budget_sheet.append([budget.high_plan.name, budget.high_plan.target_user, budget.high_plan.total_price, budget.high_plan.note])
        budget_sheet.append(["差价说明", "", "", budget.difference_summary])

        doc_sheet = workbook.create_sheet("设计方案")
        doc_sheet.append(["方案标题", plan.title])
        doc_sheet.append(["设计理念", plan.concept_summary])
        doc_sheet.append(["整体风格", plan.style_description])
        doc_sheet.append(["适合人群", plan.target_users])
        doc_sheet.append(["合规声明", plan.compliance_note])

        self._format(workbook)
        filename = f"home_design_purchase_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        path = self.settings.exports_dir / filename
        workbook.save(path)
        return path

    @staticmethod
    def _format(workbook: Workbook) -> None:
        fill = PatternFill("solid", fgColor="222222")
        header_font = Font(color="FFFFFF", bold=True)
        for sheet in workbook.worksheets:
            for cell in sheet[1]:
                cell.fill = fill
                cell.font = header_font
                cell.alignment = Alignment(vertical="center")
            for column_cells in sheet.columns:
                length = max(len(str(cell.value or "")) for cell in column_cells)
                sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(length + 2, 12), 42)
            for row in sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
